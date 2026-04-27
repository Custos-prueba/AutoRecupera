# import fitz  # PyMuPDF
# import pikepdf
# import ollama
# import json
# import os
# import sys
# import base64
# from PIL import Image
# import io
# import time
# import logging

# MODELO_TEXTO = "qwen2.5:14b"
# MODELO_VISION = "qwen2.5vl:7b"

# # --- Logging ---
# logging.basicConfig(
#     level=logging.INFO,
#     format="%(asctime)s [%(levelname)s] %(message)s",
#     handlers=[
#         logging.StreamHandler(),
#         logging.FileHandler("autorecupera.log", encoding="utf-8")
#     ]
# )
# log = logging.getLogger(__name__)


# def cronometro(nombre):
#     """Context manager para medir tiempo de operaciones."""
#     class Timer:
#         def __enter__(self):
#             self.inicio = time.time()
#             log.info(f"INICIO: {nombre}")
#             return self
#         def __exit__(self, *args):
#             self.elapsed = time.time() - self.inicio
#             log.info(f"FIN: {nombre} — {self.elapsed:.2f}s")
#     return Timer()


# def unlock_pdf(input_path, output_path, password):
#     try:
#         with pikepdf.open(input_path, password=password) as pdf:
#             pdf.save(output_path)
#             return True
#     except Exception as e:
#         log.error(f"Error al desbloquear PDF: {e}")
#         return False


# def extraer_texto_por_bloques(pdf_path, bloque_chars=2000):
#     doc = fitz.open(pdf_path)
#     bloques = []
#     texto_acumulado = ""
#     for page in doc:
#         texto_acumulado += page.get_text()
#         if len(texto_acumulado) >= bloque_chars:
#             bloques.append(texto_acumulado[:bloque_chars])
#             texto_acumulado = texto_acumulado[bloque_chars:]
#     if texto_acumulado.strip():
#         bloques.append(texto_acumulado)
#     doc.close()
#     return bloques


# def redimensionar_imagen(img_bytes, max_px=512):
#     img = Image.open(io.BytesIO(img_bytes))
#     img.thumbnail((max_px, max_px))
#     buf = io.BytesIO()
#     img.save(buf, format="JPEG", quality=75)
#     return buf.getvalue()


# def pdf_to_images(pdf_path, max_imagenes=6, min_ancho=300, min_alto=200):
#     doc = fitz.open(pdf_path)
#     imagenes = []
#     for page in doc:
#         for img in page.get_images():
#             if len(imagenes) >= max_imagenes:
#                 break
#             xref = img[0]
#             base_image = doc.extract_image(xref)
#             ancho = base_image.get("width", 0)
#             alto = base_image.get("height", 0)
#             if ancho < min_ancho or alto < min_alto:
#                 log.debug(f"Imagen ignorada ({ancho}x{alto}) — demasiado pequeña")
#                 continue
#             img_bytes = redimensionar_imagen(base_image["image"], max_px=512)
#             img_b64 = base64.b64encode(img_bytes).decode("utf-8")
#             imagenes.append(img_b64)
#             log.info(f"Imagen aceptada ({ancho}x{alto})")
#         if len(imagenes) >= max_imagenes:
#             break
#     doc.close()
#     log.info(f"Imagenes extraidas: {len(imagenes)}")
#     return imagenes


# def extraer_bloque(texto, campos):
#     lista_campos = "\n".join(f"{c}:" for c in campos)
#     prompt = f"""Extrae del siguiente texto estos datos, uno por linea con formato CLAVE: VALOR.
# Si no encuentras el dato escribe CLAVE: null
# No añadas explicaciones ni texto extra.

# {lista_campos}

# IMPORTANTE:
# - compromiso_pago debe ser "S" o "N"
# - matricula es la matricula del vehiculo (ej: 3486MFC), NO el bastidor
# - bastidor es el numero de chasis VIN (empieza por WB, longitud 17 caracteres)
# - subtotal_piezas es solo el coste de repuestos (sin mano de obra)
# - total_mo_chapa es solo la mano de obra
# - acabado es el tipo de pintura (ej: BICAPA, UNICAPA)
# - equipamiento son los extras del vehiculo (ej: PARABRISAS ALTO, TOP CASE)
# - codigo_audatransfer es un codigo hexadecimal de 6 caracteres (ej: 88D5FB)
# - fecha_siniestro es la fecha del accidente, diferente a la fecha del informe
# - total_sin_iva es "SUMA TOTAL SIN IVA" en el documento
# - total_con_iva es "SUMA TOTAL CON IVA" en el documento
# - iva_pct es el porcentaje de IVA aplicado (ej: 21)

# Texto:
# {texto}
# """
#     response = ollama.chat(
#         model=MODELO_TEXTO,
#         messages=[{"role": "user", "content": prompt}],
#         keep_alive=0,
#     )
#     lineas = response.message.content.strip().splitlines()
#     campos_extraidos = {}
#     for linea in lineas:
#         if ":" in linea:
#             clave, _, valor = linea.partition(":")
#             valor = valor.strip()
#             campos_extraidos[clave.strip()] = None if valor.lower() in ("null", "") else valor
#     return campos_extraidos


# def extraer_datos_texto(bloques):
#     campos_bloque = [
#         ["nr_informe", "referencia", "fecha_informe", "fecha_siniestro",
#          "fabricante", "modelo", "matricula", "bastidor", "compromiso_pago",
#          "codigo_audatransfer"],
#         ["fecha_matriculacion", "km", "cilindrada_cc", "potencia_cv", "potencia_kw",
#          "color", "acabado", "equipamiento", "codigo_tipo", "fecha_tarifa_recambios"],
#         ["subtotal_piezas", "descuentos", "total_mo_chapa", "horas_mo",
#          "precio_hora_chapa", "precio_hora_pintura", "total_sin_iva", "iva_pct",
#          "total_con_iva"],
#         ["total_sin_iva", "iva_pct", "total_con_iva", "descuentos",
#          "horas_mo", "precio_hora_pintura"],
#     ]

#     campos_acumulados = {}
#     for i, bloque in enumerate(bloques[:4]):
#         if i < len(campos_bloque):
#             with cronometro(f"Bloque {i+1} — modelo texto"):
#                 log.info(f"Procesando bloque {i+1} ({len(bloque)} chars)...")
#                 resultado = extraer_bloque(bloque, campos_bloque[i])
#                 for k, v in resultado.items():
#                     if v is not None and campos_acumulados.get(k) is None:
#                         campos_acumulados[k] = v

#     c = campos_acumulados
#     return {
#         "informe": {
#             "nr_informe": c.get("nr_informe"),
#             "referencia": c.get("referencia"),
#             "fecha_informe": c.get("fecha_informe"),
#             "fecha_siniestro": c.get("fecha_siniestro"),
#             "fecha_tarifa_recambios": c.get("fecha_tarifa_recambios"),
#             "codigo_audatransfer": c.get("codigo_audatransfer"),
#             "compromiso_pago": c.get("compromiso_pago"),
#         },
#         "vehiculo": {
#             "fabricante": c.get("fabricante"),
#             "modelo": c.get("modelo"),
#             "matricula": c.get("matricula"),
#             "bastidor": c.get("bastidor"),
#             "codigo_tipo": c.get("codigo_tipo"),
#             "fecha_matriculacion": c.get("fecha_matriculacion"),
#             "km": c.get("km"),
#             "cilindrada_cc": c.get("cilindrada_cc"),
#             "potencia_cv": c.get("potencia_cv"),
#             "potencia_kw": c.get("potencia_kw"),
#             "color": c.get("color"),
#             "acabado": c.get("acabado"),
#             "equipamiento": c.get("equipamiento"),
#         },
#         "valoracion": {
#             "subtotal_piezas": c.get("subtotal_piezas"),
#             "descuentos": c.get("descuentos"),
#             "total_mo_chapa": c.get("total_mo_chapa"),
#             "horas_mo": c.get("horas_mo"),
#             "precio_hora_chapa": c.get("precio_hora_chapa"),
#             "precio_hora_pintura": c.get("precio_hora_pintura"),
#             "total_sin_iva": c.get("total_sin_iva"),
#             "iva_pct": c.get("iva_pct"),
#             "total_con_iva": c.get("total_con_iva"),
#         },
#         "piezas_sustituidas": [],
#         "operaciones_mo": []
#     }


# def describir_imagenes(imagenes, datos_ya_extraidos):
#     """Usa qwen2.5vl — solo busca en imágenes lo que NO se encontró en el texto."""
#     if not imagenes:
#         return "Sin imagenes de daños disponibles."

#     # Campos que ya tenemos del texto — no hace falta buscarlos en imágenes
#     campos_encontrados = []
#     v = datos_ya_extraidos.get("vehiculo", {})
#     inf = datos_ya_extraidos.get("informe", {})
#     if v.get("km"):
#         campos_encontrados.append(f"kilometros ({v['km']} km)")
#     if v.get("matricula"):
#         campos_encontrados.append(f"matricula ({v['matricula']})")
#     if v.get("bastidor"):
#         campos_encontrados.append(f"bastidor ({v['bastidor']})")
#     if inf.get("fecha_siniestro"):
#         campos_encontrados.append(f"fecha siniestro ({inf['fecha_siniestro']})")

#     ya_tenemos = ""
#     if campos_encontrados:
#         ya_tenemos = f"\nNOTA: Ya tenemos del texto — {', '.join(campos_encontrados)}. No hace falta buscarlos en las imagenes.\n"

#     prompt = f"""Eres un perito de seguros analizando fotos de un vehiculo siniestrado.
# De todas las imagenes que ves, identifica cuales muestran daños fisicos al vehiculo.
# Para cada imagen con daños describe la zona afectada, tipo y gravedad del daño.
# {ya_tenemos}
# Si ves datos tecnicos que NO tenemos aun (matricula, VIN, kilometros), extráelos.
# Ignora imagenes que solo muestren documentos, sellos o texto del PDF.
# Responde en español en texto plano, sin asteriscos, sin markdown, sin formato especial.
# Se directo y conciso."""

#     response = ollama.chat(
#         model=MODELO_VISION,
#         messages=[{
#             "role": "user",
#             "content": prompt,
#             "images": imagenes,
#         }],
#         keep_alive=0,
#     )
#     return response.message.content


# def procesar_pdf(archivo_pdf, password=None):
#     temp_unlocked = "temp_unlocked.pdf"
#     inicio_total = time.time()
#     log.info(f"=== INICIO PROCESAMIENTO: {archivo_pdf} ===")

#     if password:
#         log.info("Desbloqueando PDF con contrasena...")
#         if not unlock_pdf(archivo_pdf, temp_unlocked, password):
#             return {"error": "No se pudo desbloquear el PDF"}
#         path_to_process = temp_unlocked
#     else:
#         path_to_process = archivo_pdf

#     try:
#         with cronometro("Extraccion de texto"):
#             bloques = extraer_texto_por_bloques(path_to_process, bloque_chars=2000)
#             log.info(f"{len(bloques)} bloque(s) de texto encontrados.")

#         if not bloques or len("".join(bloques).strip()) < 50:
#             return {"error": "PDF escaneado o vacio - se necesita OCR"}

#         with cronometro("Analisis de texto completo"):
#             data = extraer_datos_texto(bloques)
#         del bloques

#         with cronometro("Extraccion de imagenes"):
#             imagenes = pdf_to_images(path_to_process, max_imagenes=6, min_ancho=300, min_alto=200)

#         with cronometro("Descripcion visual de daños"):
#             descripcion = describir_imagenes(imagenes, data)
#         del imagenes

#         data["descripcion_visual_danos"] = descripcion

#         total = time.time() - inicio_total
#         log.info(f"=== FIN PROCESAMIENTO — total: {total:.2f}s ===")
#         return data

#     except Exception as e:
#         log.error(f"Error: {e}")
#         return {"error": str(e)}
#     finally:
#         if os.path.exists(temp_unlocked):
#             os.remove(temp_unlocked)


# if __name__ == "__main__":
#     if len(sys.argv) < 2:
#         print("Uso: python prueba_ollama.py <archivo.pdf> [contrasena]")
#         print("Ejemplo: python prueba_ollama.py informe_audatex.pdf")
#         print("Ejemplo: python prueba_ollama.py informe_audatex.pdf mipassword123")
#         sys.exit(1)

#     pdf = sys.argv[1]
#     password = sys.argv[2] if len(sys.argv) > 2 else None

#     if not os.path.exists(pdf):
#         log.error(f"No se encuentra el archivo {pdf}")
#         sys.exit(1)

#     resultado = procesar_pdf(pdf, password=password)

#     print("\n--- RESULTADO ---")
#     print(json.dumps(resultado, indent=4, ensure_ascii=False))

#     salida = pdf.replace(".pdf", "_resultado.json")
#     with open(salida, "w", encoding="utf-8") as f:
#         json.dump(resultado, f, indent=4, ensure_ascii=False)
#     log.info(f"Guardado en: {salida}")


import pypdf
import pikepdf
import ollama
import json
import os
import sys
import base64
from PIL import Image
import io
import time
import logging
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

MODELO_TEXTO = "qwen2.5:14b"
MODELO_VISION = "qwen2.5vl:7b"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler("autorecupera.log", encoding="utf-8")
    ]
)
log = logging.getLogger(__name__)

COLOR_CABECERA    = "1F4E79"
COLOR_SECCION     = "2E75B6"
COLOR_FILA_PAR    = "DEEAF1"
COLOR_TEXTO_BLANCO = "FFFFFF"
COLOR_TEXTO_NEGRO  = "000000"

def _borde_fino():
    lado = Side(style="thin", color="AAAAAA")
    return Border(left=lado, right=lado, top=lado, bottom=lado)

def _estilo_cabecera(cell, texto):
    cell.value = texto
    cell.font = Font(bold=True, color=COLOR_TEXTO_BLANCO, size=11)
    cell.fill = PatternFill("solid", start_color=COLOR_CABECERA)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _borde_fino()

def _estilo_seccion(cell, texto):
    cell.value = texto
    cell.font = Font(bold=True, color=COLOR_TEXTO_BLANCO, size=10)
    cell.fill = PatternFill("solid", start_color=COLOR_SECCION)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = _borde_fino()

def _estilo_clave(cell, texto):
    cell.value = texto
    cell.font = Font(bold=True, color=COLOR_TEXTO_NEGRO, size=10)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = _borde_fino()

def _estilo_valor(cell, valor, fila_par=False):
    cell.value = valor if valor is not None else "—"
    cell.font = Font(color=COLOR_TEXTO_NEGRO, size=10)
    cell.fill = PatternFill("solid", start_color=COLOR_FILA_PAR if fila_par else "FFFFFF")
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = _borde_fino()

def guardar_excel(data, ruta_salida, ruta_pdf):
    if os.path.exists(ruta_salida):
        wb = load_workbook(ruta_salida)
    else:
        wb = Workbook()
        wb.active.title = "Historial"

    _actualizar_historial(wb, data, ruta_pdf)
    _crear_hoja_detalle(wb, data)

    wb.save(ruta_salida)
    log.info(f"Excel guardado en: {ruta_salida}")

def _actualizar_historial(wb, data, ruta_pdf):
    if "Historial" not in wb.sheetnames:
        ws = wb.create_sheet("Historial", 0)
    else:
        ws = wb["Historial"]

    COLUMNAS = [
        "Archivo", "Nº Informe", "Referencia", "Fecha informe",
        "Fecha siniestro", "Fabricante", "Modelo", "Matrícula",
        "Bastidor", "Km", "Color",
        "Total sin IVA", "IVA %", "Total con IVA",
        "Compromiso pago", "Código Audatransfer",
    ]

    if ws.max_row == 1 and ws["A1"].value is None:
        for col, texto in enumerate(COLUMNAS, start=1):
            _estilo_cabecera(ws.cell(row=1, column=col), texto)
        ws.row_dimensions[1].height = 22

    inf = data.get("informe", {})
    veh = data.get("vehiculo", {})
    val = data.get("valoracion", {})

    fila = ws.max_row + 1
    valores = [
        os.path.basename(ruta_pdf),
        inf.get("nr_informe"),
        inf.get("referencia"),
        inf.get("fecha_informe"),
        inf.get("fecha_siniestro"),
        veh.get("fabricante"),
        veh.get("modelo"),
        veh.get("matricula"),
        veh.get("bastidor"),
        veh.get("km"),
        veh.get("color"),
        val.get("total_sin_iva"),
        val.get("iva_pct"),
        val.get("total_con_iva"),
        inf.get("compromiso_pago"),
        inf.get("codigo_audatransfer"),
    ]
    par = fila % 2 == 0
    for col, v in enumerate(valores, start=1):
        _estilo_valor(ws.cell(row=fila, column=col), v, fila_par=par)

    anchos = [28, 14, 16, 14, 14, 14, 18, 12, 20, 10, 12, 14, 8, 14, 16, 18]
    for col, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(col)].width = ancho

def _crear_hoja_detalle(wb, data):
    nombre = "Último informe"
    if nombre in wb.sheetnames:
        del wb[nombre]
    ws = wb.create_sheet(nombre)

    inf = data.get("informe", {})
    veh = data.get("vehiculo", {})
    val = data.get("valoracion", {})

    secciones = [
        ("DATOS DEL INFORME", [
            ("Nº informe",          inf.get("nr_informe")),
            ("Referencia",          inf.get("referencia")),
            ("Fecha informe",       inf.get("fecha_informe")),
            ("Fecha siniestro",     inf.get("fecha_siniestro")),
            ("Fecha tarifa recambios", inf.get("fecha_tarifa_recambios")),
            ("Código Audatransfer", inf.get("codigo_audatransfer")),
            ("Compromiso pago",     inf.get("compromiso_pago")),
        ]),
        ("DATOS DEL VEHÍCULO", [
            ("Fabricante",          veh.get("fabricante")),
            ("Modelo",              veh.get("modelo")),
            ("Matrícula",           veh.get("matricula")),
            ("Bastidor",            veh.get("bastidor")),
            ("Código tipo",         veh.get("codigo_tipo")),
            ("Fecha matriculación", veh.get("fecha_matriculacion")),
            ("Kilómetros",          veh.get("km")),
            ("Cilindrada (cc)",     veh.get("cilindrada_cc")),
            ("Potencia (CV)",       veh.get("potencia_cv")),
            ("Potencia (kW)",       veh.get("potencia_kw")),
            ("Color",               veh.get("color")),
            ("Acabado",             veh.get("acabado")),
            ("Equipamiento",        veh.get("equipamiento")),
        ]),
        ("VALORACIÓN ECONÓMICA", [
            ("Subtotal piezas",     val.get("subtotal_piezas")),
            ("Descuentos",          val.get("descuentos")),
            ("Total MO chapa",      val.get("total_mo_chapa")),
            ("Horas MO",            val.get("horas_mo")),
            ("Precio hora chapa",   val.get("precio_hora_chapa")),
            ("Precio hora pintura", val.get("precio_hora_pintura")),
            ("Total sin IVA",       val.get("total_sin_iva")),
            ("IVA %",               val.get("iva_pct")),
            ("Total con IVA",       val.get("total_con_iva")),
        ]),
    ]

    ws.merge_cells("A1:B1")
    _estilo_cabecera(ws["A1"], "INFORME DE PERITACIÓN — RESULTADO EXTRACCIÓN")
    ws.row_dimensions[1].height = 24
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 40

    fila = 2
    par = False
    for titulo_seccion, campos in secciones:
        ws.merge_cells(f"A{fila}:B{fila}")
        _estilo_seccion(ws[f"A{fila}"], titulo_seccion)
        ws.row_dimensions[fila].height = 20
        fila += 1
        for clave, valor in campos:
            _estilo_clave(ws.cell(row=fila, column=1), clave)
            _estilo_valor(ws.cell(row=fila, column=2), valor, fila_par=par)
            ws.row_dimensions[fila].height = 18
            par = not par
            fila += 1
        fila += 1

    desc = data.get("descripcion_visual_danos", "")
    if desc:
        ws.merge_cells(f"A{fila}:B{fila}")
        _estilo_seccion(ws[f"A{fila}"], "DESCRIPCIÓN VISUAL DE DAÑOS")
        ws.row_dimensions[fila].height = 20
        fila += 1
        ws.merge_cells(f"A{fila}:B{fila + 10}")
        celda = ws[f"A{fila}"]
        celda.value = desc
        celda.font = Font(size=10)
        celda.alignment = Alignment(wrap_text=True, vertical="top")
        celda.border = _borde_fino()
        ws.row_dimensions[fila].height = 120

def cronometro(nombre):
    class Timer:
        def __enter__(self):
            self.inicio = time.time()
            log.info(f"INICIO: {nombre}")
            return self
        def __exit__(self, *args):
            self.elapsed = time.time() - self.inicio
            log.info(f"FIN: {nombre} — {self.elapsed:.2f}s")
    return Timer()

def unlock_pdf(input_path, output_path, password):
    try:
        with pikepdf.open(input_path, password=password) as pdf:
            pdf.save(output_path)
        return True
    except Exception as e:
        log.error(f"Error al desbloquear PDF: {e}")
        return False

def extraer_texto_por_bloques(pdf_path, bloque_chars=2000):
    bloques = []
    texto_acumulado = ""
    with pypdf.PdfReader(pdf_path) as doc:
        for page in doc.pages:
            try:
                texto_acumulado += page.extract_text()
            except Exception as e:
                log.warning(f"Error extrayendo texto de página: {e}")
                continue
            if len(texto_acumulado) >= bloque_chars:
                bloques.append(texto_acumulado[:bloque_chars])
                texto_acumulado = texto_acumulado[bloque_chars:]
    if texto_acumulado.strip():
        bloques.append(texto_acumulado)
    return bloques

def redimensionar_imagen(img_bytes, max_px=512):
    img = Image.open(io.BytesIO(img_bytes))
    img.thumbnail((max_px, max_px))
    buf = io.BytesIO()
    img.save(buf, format="JPEG", quality=75)
    return buf.getvalue()

def pdf_to_images(pdf_path, max_imagenes=6, min_ancho=300, min_alto=200):
    imagenes = []
    try:
        with pypdf.PdfReader(pdf_path) as doc:
            for page_num, page in enumerate(doc.pages):
                if len(imagenes) >= max_imagenes:
                    break
                for img_index, img in enumerate(page.images):
                    if len(imagenes) >= max_imagenes:
                        break
                    try:
                        img_data = img.get_object()
                        width = img_data.get("/Width", 0)
                        height = img_data.get("/Height", 0)
                        if width < min_ancho or height < min_alto:
                            log.debug(f"Imagen ignorada ({width}x{height})")
                            continue
                        img_bytes = img.get_data()
                        img_bytes = redimensionar_imagen(img_bytes, max_px=512)
                        img_b64 = base64.b64encode(img_bytes).decode("utf-8")
                        imagenes.append(img_b64)
                        log.info(f"Imagen aceptada ({width}x{height})")
                    except Exception as e:
                        log.warning(f"Error procesando imagen: {e}")
                        continue
    except Exception as e:
        log.warning(f"Error extrayendo imágenes: {e}")
    
    log.info(f"Imagenes extraidas: {len(imagenes)}")
    return imagenes

def extraer_bloque(texto, campos):
    lista_campos = "\n".join(f"{c}:" for c in campos)
    prompt = f"""Extrae del siguiente texto estos datos, uno por linea con formato CLAVE: VALOR.
Si no encuentras el dato escribe CLAVE: null
No añadas explicaciones ni texto extra.

{lista_campos}

IMPORTANTE:
- compromiso_pago debe ser "S" o "N"
- matricula es la matricula del vehiculo (ej: 3486MFC), NO el bastidor
- bastidor es el numero de chasis VIN (empieza por WB, longitud 17 caracteres)
- subtotal_piezas es solo el coste de repuestos (sin mano de obra)
- total_mo_chapa es solo la mano de obra
- acabado es el tipo de pintura (ej: BICAPA, UNICAPA)
- equipamiento son los extras del vehiculo (ej: PARABRISAS ALTO, TOP CASE)
- codigo_audatransfer es un codigo hexadecimal de 6 caracteres (ej: 88D5FB)
- fecha_siniestro es la fecha del accidente, diferente a la fecha del informe
- total_sin_iva es "SUMA TOTAL SIN IVA" en el documento
- total_con_iva es "SUMA TOTAL CON IVA" en el documento
- iva_pct es el porcentaje de IVA aplicado (ej: 21)

Texto:
{texto}
"""
    response = ollama.chat(
        model=MODELO_TEXTO,
        messages=[{"role": "user", "content": prompt}],
        keep_alive=0,
    )
    lineas = response.message.content.strip().splitlines()
    campos_extraidos = {}
    for linea in lineas:
        if ":" in linea:
            clave, _, valor = linea.partition(":")
            valor = valor.strip()
            campos_extraidos[clave.strip()] = None if valor.lower() in ("null", "") else valor
    return campos_extraidos

def extraer_datos_texto(bloques):
    campos_bloque = [
        ["nr_informe", "referencia", "fecha_informe", "fecha_siniestro",
         "fabricante", "modelo", "matricula", "bastidor", "compromiso_pago",
         "codigo_audatransfer"],
        ["fecha_matriculacion", "km", "cilindrada_cc", "potencia_cv", "potencia_kw",
         "color", "acabado", "equipamiento", "codigo_tipo", "fecha_tarifa_recambios"],
        ["subtotal_piezas", "descuentos", "total_mo_chapa", "horas_mo",
         "precio_hora_chapa", "precio_hora_pintura", "total_sin_iva", "iva_pct",
         "total_con_iva"],
        ["total_sin_iva", "iva_pct", "total_con_iva", "descuentos",
         "horas_mo", "precio_hora_pintura"],
    ]

    campos_acumulados = {}
    for i, bloque in enumerate(bloques[:4]):
        if i < len(campos_bloque):
            with cronometro(f"Bloque {i+1}"):
                log.info(f"Procesando bloque {i+1} ({len(bloque)} chars)...")
                resultado = extraer_bloque(bloque, campos_bloque[i])
                for k, v in resultado.items():
                    if v is not None and campos_acumulados.get(k) is None:
                        campos_acumulados[k] = v

    c = campos_acumulados
    return {
        "informe": {
            "nr_informe": c.get("nr_informe"),
            "referencia": c.get("referencia"),
            "fecha_informe": c.get("fecha_informe"),
            "fecha_siniestro": c.get("fecha_siniestro"),
            "fecha_tarifa_recambios": c.get("fecha_tarifa_recambios"),
            "codigo_audatransfer": c.get("codigo_audatransfer"),
            "compromiso_pago": c.get("compromiso_pago"),
        },
        "vehiculo": {
            "fabricante": c.get("fabricante"),
            "modelo": c.get("modelo"),
            "matricula": c.get("matricula"),
            "bastidor": c.get("bastidor"),
            "codigo_tipo": c.get("codigo_tipo"),
            "fecha_matriculacion": c.get("fecha_matriculacion"),
            "km": c.get("km"),
            "cilindrada_cc": c.get("cilindrada_cc"),
            "potencia_cv": c.get("potencia_cv"),
            "potencia_kw": c.get("potencia_kw"),
            "color": c.get("color"),
            "acabado": c.get("acabado"),
            "equipamiento": c.get("equipamiento"),
        },
        "valoracion": {
            "subtotal_piezas": c.get("subtotal_piezas"),
            "descuentos": c.get("descuentos"),
            "total_mo_chapa": c.get("total_mo_chapa"),
            "horas_mo": c.get("horas_mo"),
            "precio_hora_chapa": c.get("precio_hora_chapa"),
            "precio_hora_pintura": c.get("precio_hora_pintura"),
            "total_sin_iva": c.get("total_sin_iva"),
            "iva_pct": c.get("iva_pct"),
            "total_con_iva": c.get("total_con_iva"),
        },
        "piezas_sustituidas": [],
        "operaciones_mo": []
    }

def describir_imagenes(imagenes, datos_ya_extraidos):
    if not imagenes:
        return "Sin imagenes de daños disponibles."

    campos_encontrados = []
    v = datos_ya_extraidos.get("vehiculo", {})
    inf = datos_ya_extraidos.get("informe", {})
    if v.get("km"):
        campos_encontrados.append(f"kilometros ({v['km']} km)")
    if v.get("matricula"):
        campos_encontrados.append(f"matricula ({v['matricula']})")
    if v.get("bastidor"):
        campos_encontrados.append(f"bastidor ({v['bastidor']})")
    if inf.get("fecha_siniestro"):
        campos_encontrados.append(f"fecha siniestro ({inf['fecha_siniestro']})")

    ya_tenemos = ""
    if campos_encontrados:
        ya_tenemos = f"\nNOTA: Ya tenemos del texto — {', '.join(campos_encontrados)}. No hace falta buscarlos en las imagenes.\n"

    prompt = f"""Eres un perito de seguros analizando fotos de un vehiculo siniestrado.
De todas las imagenes que ves, identifica cuales muestran daños fisicos al vehiculo.
Para cada imagen con daños describe la zona afectada, tipo y gravedad del daño.
{ya_tenemos}
Si ves datos tecnicos que NO tenemos aun, extráelos.
Ignora imagenes que solo muestren documentos, sellos o texto del PDF.
Responde en español en texto plano, sin asteriscos, sin markdown, sin formato especial.
Se directo y conciso."""

    response = ollama.chat(
        model=MODELO_VISION,
        messages=[{
            "role": "user",
            "content": prompt,
            "images": imagenes,
        }],
        keep_alive=0,
    )
    return response.message.content

def procesar_pdf(archivo_pdf, password=None):
    temp_unlocked = "temp_unlocked.pdf"
    inicio_total = time.time()
    log.info(f"=== INICIO PROCESAMIENTO: {archivo_pdf} ===")

    if password:
        if not unlock_pdf(archivo_pdf, temp_unlocked, password):
            return {"error": "No se pudo desbloquear el PDF"}
        path_to_process = temp_unlocked
    else:
        path_to_process = archivo_pdf

    try:
        with cronometro("Extraccion de texto"):
            bloques = extraer_texto_por_bloques(path_to_process, bloque_chars=2000)
            log.info(f"{len(bloques)} bloque(s) encontrados")

        if not bloques or len("".join(bloques).strip()) < 50:
            return {"error": "PDF escaneado o vacio"}

        with cronometro("Analisis de texto"):
            data = extraer_datos_texto(bloques)
        del bloques

        with cronometro("Extraccion de imagenes"):
            imagenes = pdf_to_images(path_to_process, max_imagenes=6, min_ancho=300, min_alto=200)

        with cronometro("Descripcion de daños"):
            descripcion = describir_imagenes(imagenes, data)
        del imagenes

        data["descripcion_visual_danos"] = descripcion

        total = time.time() - inicio_total
        log.info(f"=== FIN PROCESAMIENTO — total: {total:.2f}s ===")
        return data

    except Exception as e:
        log.error(f"Error: {e}")
        return {"error": str(e)}
    finally:
        if os.path.exists(temp_unlocked):
            os.remove(temp_unlocked)

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python autorecupera_freebsd.py <archivo.pdf> [contrasena] [salida.xlsx]")
        print("\nEjemplos:")
        print("  python autorecupera_freebsd.py informe.pdf")
        print("  python autorecupera_freebsd.py informe.pdf mipass resultados.xlsx")
        sys.exit(1)

    pdf      = sys.argv[1]
    password = sys.argv[2] if len(sys.argv) > 2 and sys.argv[2] else None
    salida   = sys.argv[3] if len(sys.argv) > 3 else "resultados_peritos.xlsx"

    if not os.path.exists(pdf):
        log.error(f"No se encuentra: {pdf}")
        sys.exit(1)

    resultado = procesar_pdf(pdf, password=password)

    if "error" in resultado:
        log.error(f"Error: {resultado['error']}")
        sys.exit(1)

    guardar_excel(resultado, salida, pdf)
    print(f"\nExcel guardado en: {salida}")